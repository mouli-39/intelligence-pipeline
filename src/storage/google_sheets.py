import os
import csv
from typing import List, Dict, Any
from src.utils.logging import setup_logger

logger = setup_logger("storage_engine")

class GoogleSheetsDataStore:
    """Asynchronous-safe storage engine structuring pipeline records into multi-tab sheets formats."""
    
    def __init__(self, base_output_dir: str = "data/processed", excel_output_path: str = "data/AI_Intelligence_Pipeline_Sheet.xlsx"):
        self.output_dir = base_output_dir
        self.excel_path = excel_output_path
        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(os.path.dirname(self.excel_path) or ".", exist_ok=True)
        
        # Define uniform file locations matching the 6 requested tabs
        self.tab_mappings = {
            "Startups": os.path.join(self.output_dir, "tab_startups.csv"),
            "Products": os.path.join(self.output_dir, "tab_products.csv"),
            "Research Papers": os.path.join(self.output_dir, "tab_research_papers.csv"),
            "Jobs": os.path.join(self.output_dir, "tab_jobs.csv"),
            "News": os.path.join(self.output_dir, "tab_news.csv"),
            "Entity Mapping Log": os.path.join(self.output_dir, "tab_entity_mapping_log.csv")
        }
        self._initialize_headers()

    def _initialize_headers(self):
        """Pre-seeds standard data column headers across all tabs safely."""
        headers = {
            "Startups": ["Entity Name", "Employee Count", "Source Name", "Source URL", "Collected At"],
            "Products": ["Product Name", "Pricing Model", "Source Name", "Source URL", "Collected At"],
            "Research Papers": ["Title", "Authors", "Paper URL", "GitHub URL", "GitHub Stars", "Published Date"],
            "Jobs": ["Job Title", "Company", "Source Name", "Source URL", "Date Collected"],
            "News": ["Article Title", "Source Channel", "Source URL", "Published Date"],
            "Entity Mapping Log": ["Raw Input Value", "Resolved Canonical Entity"]
        }
        
        for tab_name, file_path in self.tab_mappings.items():
            if not os.path.exists(file_path):
                with open(file_path, mode="w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(headers[tab_name])
                logger.info(f"Initialized storage sheet partition: {tab_name}")

    def append_rows_to_tab(self, tab_name: str, rows: List[List[Any]]):
        """Appends streaming processed records immediately to a specific data sheet tab."""
        if tab_name not in self.tab_mappings:
            raise ValueError(f"Target tab '{tab_name}' is not supported by the pipeline schema.")
            
        file_path = self.tab_mappings[tab_name]
        try:
            with open(file_path, mode="a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerows(rows)
            logger.info(f"Successfully appended {len(rows)} records to Sheet Tab: {tab_name}")
        except Exception as e:
            logger.error(f"Failed to commit row batch operation to {tab_name}: {str(e)}")

    def export_unified_spreadsheet(self):
        """Compiles CSV partitions into a single multi-tab Excel/Google Sheets compatible workbook."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            # Remove default active sheet
            default_sheet = wb.active
            
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            
            for index, (tab_name, csv_path) in enumerate(self.tab_mappings.items()):
                ws = wb.create_sheet(title=tab_name)
                
                if os.path.exists(csv_path):
                    with open(csv_path, mode="r", encoding="utf-8") as f:
                        reader = csv.reader(f)
                        for row_idx, row in enumerate(reader, 1):
                            ws.append(row)
                            if row_idx == 1:
                                for cell in ws[1]:
                                    cell.fill = header_fill
                                    cell.font = header_font
                                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-adjust column widths
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
            if default_sheet in wb.worksheets:
                wb.remove(default_sheet)
                
            wb.save(self.excel_path)
            logger.info(f"Unified multi-tab spreadsheet created successfully at: {self.excel_path}")
        except Exception as e:
            logger.error(f"Failed to export multi-tab spreadsheet: {str(e)}")

